import pandas as pd
from openpyxl import load_workbook
from config import EXCEL_FILE, DATE_LOG


def init_log():
    try:
        log_df = pd.read_excel(DATE_LOG)
    except FileNotFoundError:
        log_df = pd.DataFrame(columns=[
            "Задача", 
            "Дата и время напоминания", 
            "Дата получения ответа", 
            "Email"
        ])
    return log_df


def column_width(file_path):
    # Путь к файлу
    file_path = DATE_LOG

    # 1. Загружаем данные в Pandas
    df = pd.read_excel(file_path)

# 2. Производим нужные изменения в DataFrame
# Например:
# df["Новая колонка"] = df["Старая колонка"] * 2

# 3. Загружаем книгу Excel, чтобы сохранить форматирование
    book = load_workbook(file_path)

# 4. Получаем ширины колонок из исходного листа (если нужно сохранить)
    sheet_name = "Sheet1"  # укажите ваш лист
    sheet = book[sheet_name]

# Сохраняем ширины колонок (если важно)
    column_widths = {}
    for col in sheet.columns:
        col_letter = col[0].column_letter
        column_widths[col_letter] = sheet.column_dimensions[col_letter].width

# 5. Удаляем старый лист (если нужно)
    if sheet_name in book.sheetnames:
        del book[sheet_name]

# 6. Сохраняем DataFrame в тот же файл
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# 7. Заново открываем книгу и восстанавливаем ширины колонок
    book = load_workbook(file_path)
    sheet = book[sheet_name]

# Применяем сохраненные ширины
    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

# 8. Сохраняем изменения
    book.save(file_path)

def load_tasks():
    try:
        return pd.read_excel(EXCEL_FILE)
    except FileNotFoundError:
        print("Файл tasks.xlsx не найден, создаю новый...")
        df = pd.DataFrame(columns=["Задача", "Исполнитель", "Email", "Дедлайн", "Статус"])
        df.to_excel(EXCEL_FILE, index=False)
        return df
    

def update_status(email, status):
    try:
        df = pd.read_excel(EXCEL_FILE)
        # Ищем точное совпадение email (без учета регистра)
        mask = df["Email"].str.lower() == email.lower()
        if not any(mask):
            print(f" Email {email} не найден в tasks.xlsx")
            return
            
        df.loc[mask, "Статус"] = status
        df.to_excel(EXCEL_FILE, index=False)
        print(f" Обновлен статус для {email}: {status}")
    except Exception as e:
        print(f" Ошибка при обновлении статуса: {e}")